"""
generate_summary.py
-------------------
Produces a benchmark results Excel workbook with multiple sheets:
  - Sheet 1 "TDC_Leaderboard": Official TDC top 5 per task
  - Sheet 2+ per model: Reproduced results across tasks
  - Also saves benchmark_comparison.csv for quick reference

Usage:
    python generate_summary.py
    python generate_summary.py --logs ./model_assets --output ./results
"""

import os
import json
import argparse
import pandas as pd
from glob import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MODELS = [
    "MiniMol",
    "DeepMol",
    "MapLight_GNN",
    "AttrMasking",
    "AttentiveFP",
    "BasicML",
    "CaliciBoost",
    "BaseBoosting",
    "ZairaChem",
]

MODEL_ALGORITHM = {
    "MiniMol":      "Random Forest on GNN fingerprint",
    "DeepMol":      "Random Forest / XGBoost (AutoML)",
    "MapLight_GNN": "CatBoost on concatenated fingerprints + GIN embeddings",
    "AttrMasking":  "Fine-tuned GIN (pre-trained, attribute masking)",
    "AttentiveFP":  "Graph Attention Network (trains from scratch)",
    "BasicML":      "Best of 10 sklearn/XGBoost models on 31 RDKit descriptors",
    "CaliciBoost":  "XGBoost on top-ranked PaDEL descriptors (Caco-2 only)",
    "BaseBoosting": "Gradient-boosted ensemble of 3 RFs (Morgan + RDKit2D + GIN)",
    "ZairaChem":    "Ensemble (XGBoost/LightGBM/RF/NN on Ersilia descriptors)",
}

# Models that could not be run — shown as blocked sheets in Excel
BLOCKED_MODELS = {
    "ZairaChem": (
        "ZairaChem — Could Not Run",
        [
            "Reason: Ersilia Model Hub requires network access to GitHub and DockerHub.",
            "Both are blocked on the company server.",
            "eos2gw4 and eos2db3 model weights are stored in Git LFS which is also blocked.",
            "Only eos5axz (deterministic Morgan fingerprints) could be fetched.",
            "ZairaChem can only be run in an environment with unrestricted internet access.",
        ]
    ),
    "BaseBoosting": (
        "BaseBoosting — Could Not Run",
        [
            "Reason: olorenchemengine (OCE) is not available as a prebuilt wheel on PyPI.",
            "Installation requires cloning from GitHub (github.com/Oloren-AI/olorenchemengine),",
            "which is blocked by the company server firewall.",
            "All three install methods attempted: pip, Tsinghua mirror, git clone — all failed.",
            "BaseBoosting can only be run on a machine with unrestricted internet access to GitHub.",
        ]
    ),
}

# CaliciBoost only covers caco2_wang
MODEL_TASK_COVERAGE = {
    "CaliciBoost": {"caco2_wang"},  # only task this model covers
}

TASKS = [
    "caco2_wang", "hia_hou", "pgp_broccatelli", "bioavailability_ma",
    "lipophilicity_astrazeneca", "solubility_aqsoldb", "bbb_martins",
    "ppbr_az", "vdss_lombardo", "cyp2d6_veith", "cyp3a4_veith",
    "cyp2c9_veith", "cyp2c9_substrate_carbonmangels",
    "cyp2d6_substrate_carbonmangels", "cyp3a4_substrate_carbonmangels",
    "half_life_obach", "clearance_hepatocyte_az", "clearance_microsome_az",
    "ld50_zhu", "herg", "ames", "dili",
]

LOWER_IS_BETTER = {
    "caco2_wang", "lipophilicity_astrazeneca", "solubility_aqsoldb",
    "ppbr_az", "half_life_obach", "clearance_hepatocyte_az",
    "clearance_microsome_az", "ld50_zhu",
}

TASK_TYPE = {
    "caco2_wang": "regression", "hia_hou": "binary",
    "pgp_broccatelli": "binary", "bioavailability_ma": "binary",
    "lipophilicity_astrazeneca": "regression", "solubility_aqsoldb": "regression",
    "bbb_martins": "binary", "ppbr_az": "regression", "vdss_lombardo": "regression",
    "cyp2d6_veith": "binary", "cyp3a4_veith": "binary", "cyp2c9_veith": "binary",
    "cyp2c9_substrate_carbonmangels": "binary",
    "cyp2d6_substrate_carbonmangels": "binary",
    "cyp3a4_substrate_carbonmangels": "binary",
    "half_life_obach": "regression", "clearance_hepatocyte_az": "regression",
    "clearance_microsome_az": "regression", "ld50_zhu": "regression",
    "herg": "binary", "ames": "binary", "dili": "binary",
}

# Official TDC leaderboard top 5 per task
TDC_LEADERBOARD = [
    ("caco2_wang","CaliciBoost","regression","MAE",0.256,0.006),
    ("caco2_wang","XG Boost","regression","MAE",0.274,0.004),
    ("caco2_wang","MapLight","regression","MAE",0.276,0.005),
    ("caco2_wang","BaseBoosting","regression","MAE",0.285,0.005),
    ("caco2_wang","MolMapNet-D","regression","MAE",0.287,0.005),
    ("hia_hou","MiniMol","binary","AUROC",0.993,0.005),
    ("hia_hou","DeepMol (AutoML)","binary","AUROC",0.990,0.002),
    ("hia_hou","MapLight + GNN","binary","AUROC",0.989,0.001),
    ("hia_hou","RFStacker","binary","AUROC",0.988,0.002),
    ("hia_hou","MapLight","binary","AUROC",0.986,0.000),
    ("pgp_broccatelli","MapLight + GNN","binary","AUROC",0.938,0.002),
    ("pgp_broccatelli","ZairaChem","binary","AUROC",0.935,0.006),
    ("pgp_broccatelli","MapLight","binary","AUROC",0.930,0.002),
    ("pgp_broccatelli","SimGCN","binary","AUROC",0.929,0.010),
    ("pgp_broccatelli","AttrMasking","binary","AUROC",0.929,0.006),
    ("bioavailability_ma","MiniMol","binary","AUROC",0.942,0.002),
    ("bioavailability_ma","MapLight + GNN","binary","AUROC",0.938,0.002),
    ("bioavailability_ma","ZairaChem","binary","AUROC",0.935,0.006),
    ("bioavailability_ma","MapLight","binary","AUROC",0.930,0.002),
    ("bioavailability_ma","SimGCN","binary","AUROC",0.929,0.010),
    ("lipophilicity_astrazeneca","MiniMol","regression","MAE",0.456,0.008),
    ("lipophilicity_astrazeneca","Chemprop-RDKit","regression","MAE",0.467,0.006),
    ("lipophilicity_astrazeneca","Chemprop","regression","MAE",0.470,0.009),
    ("lipophilicity_astrazeneca","BaseBoosting","regression","MAE",0.479,0.007),
    ("lipophilicity_astrazeneca","ADMETrix","regression","MAE",0.512,0.010),
    ("solubility_aqsoldb","MiniMol","regression","MAE",0.741,0.013),
    ("solubility_aqsoldb","Chemprop-RDKit","regression","MAE",0.761,0.025),
    ("solubility_aqsoldb","DeepMol (AutoML)","regression","MAE",0.775,0.006),
    ("solubility_aqsoldb","AttentiveFP","regression","MAE",0.776,0.008),
    ("solubility_aqsoldb","MapLight + GNN","regression","MAE",0.789,0.003),
    ("bbb_martins","MiniMol","binary","AUROC",0.924,0.003),
    ("bbb_martins","CFA","binary","AUROC",0.920,0.006),
    ("bbb_martins","MapLight","binary","AUROC",0.916,0.001),
    ("bbb_martins","Lantern RADR Ensemble","binary","AUROC",0.915,0.002),
    ("bbb_martins","MapLight + GNN","binary","AUROC",0.913,0.001),
    ("ppbr_az","Gradient Boost","regression","MAE",7.440,0.024),
    ("ppbr_az","MapLight + GNN","regression","MAE",7.526,0.106),
    ("ppbr_az","MapLight","regression","MAE",7.660,0.058),
    ("ppbr_az","MiniMol","regression","MAE",7.696,0.125),
    ("ppbr_az","Chemprop","regression","MAE",7.788,0.210),
    ("vdss_lombardo","MapLight + GNN","regression","Spearman",0.713,0.007),
    ("vdss_lombardo","MapLight","regression","Spearman",0.707,0.009),
    ("vdss_lombardo","CFA","regression","Spearman",0.628,0.023),
    ("vdss_lombardo","Basic ML","regression","Spearman",0.627,0.010),
    ("vdss_lombardo","Euclia ML model","regression","Spearman",0.609,0.014),
    ("cyp2c9_veith","MapLight + GNN","binary","AUPRC",0.859,0.001),
    ("cyp2c9_veith","ContextPred","binary","AUPRC",0.839,0.003),
    ("cyp2c9_veith","AttrMasking","binary","AUPRC",0.829,0.003),
    ("cyp2c9_veith","MiniMol","binary","AUPRC",0.823,0.006),
    ("cyp2c9_veith","ADMETrix","binary","AUPRC",0.789,0.004),
    ("cyp2d6_veith","MapLight + GNN","binary","AUPRC",0.790,0.001),
    ("cyp2d6_veith","ContextPred","binary","AUPRC",0.739,0.005),
    ("cyp2d6_veith","MapLight","binary","AUPRC",0.723,0.003),
    ("cyp2d6_veith","AttrMasking","binary","AUPRC",0.721,0.009),
    ("cyp2d6_veith","MiniMol","binary","AUPRC",0.719,0.004),
    ("cyp3a4_veith","MapLight + GNN","binary","AUPRC",0.916,0.000),
    ("cyp3a4_veith","ContextPred","binary","AUPRC",0.904,0.002),
    ("cyp3a4_veith","AttrMasking","binary","AUPRC",0.902,0.002),
    ("cyp3a4_veith","ADMETrix","binary","AUPRC",0.884,0.001),
    ("cyp3a4_veith","MapLight","binary","AUPRC",0.881,0.001),
    ("cyp2c9_substrate_carbonmangels","MiniMol","binary","AUPRC",0.474,0.025),
    ("cyp2c9_substrate_carbonmangels","ZairaChem","binary","AUPRC",0.441,0.033),
    ("cyp2c9_substrate_carbonmangels","MapLight + GNN","binary","AUPRC",0.437,0.008),
    ("cyp2c9_substrate_carbonmangels","Random Forest","binary","AUPRC",0.437,0.022),
    ("cyp2c9_substrate_carbonmangels","SimGCN","binary","AUPRC",0.433,0.017),
    ("cyp2d6_substrate_carbonmangels","ContextPred","binary","AUPRC",0.736,0.024),
    ("cyp2d6_substrate_carbonmangels","DeepMol (AutoML)","binary","AUPRC",0.731,0.037),
    ("cyp2d6_substrate_carbonmangels","MapLight + GNN","binary","AUPRC",0.720,0.002),
    ("cyp2d6_substrate_carbonmangels","MapLight","binary","AUPRC",0.713,0.009),
    ("cyp2d6_substrate_carbonmangels","CFA","binary","AUPRC",0.704,0.015),
    ("cyp3a4_substrate_carbonmangels","CFA","binary","AUROC",0.667,0.019),
    ("cyp3a4_substrate_carbonmangels","MiniMol","binary","AUROC",0.663,0.008),
    ("cyp3a4_substrate_carbonmangels","CNN (DeepPurpose)","binary","AUROC",0.662,0.031),
    ("cyp3a4_substrate_carbonmangels","DeepMol (AutoML)","binary","AUROC",0.655,0.003),
    ("cyp3a4_substrate_carbonmangels","MapLight","binary","AUROC",0.650,0.006),
    ("half_life_obach","CFA","regression","Spearman",0.576,0.025),
    ("half_life_obach","MapLight","regression","Spearman",0.562,0.008),
    ("half_life_obach","MapLight + GNN","regression","Spearman",0.557,0.034),
    ("half_life_obach","Euclia ML model","regression","Spearman",0.547,0.032),
    ("half_life_obach","Voting Regressor","regression","Spearman",0.544,0.034),
    ("clearance_hepatocyte_az","CFA","regression","Spearman",0.536,0.020),
    ("clearance_hepatocyte_az","MapLight + GNN","regression","Spearman",0.498,0.009),
    ("clearance_hepatocyte_az","MapLight","regression","Spearman",0.466,0.012),
    ("clearance_hepatocyte_az","ADMETrix","regression","Spearman",0.447,0.028),
    ("clearance_hepatocyte_az","MiniMol","regression","Spearman",0.446,0.029),
    ("clearance_microsome_az","MapLight + GNN","regression","Spearman",0.630,0.010),
    ("clearance_microsome_az","MiniMol","regression","Spearman",0.628,0.005),
    ("clearance_microsome_az","MapLight","regression","Spearman",0.626,0.008),
    ("clearance_microsome_az","CFA","regression","Spearman",0.625,0.012),
    ("clearance_microsome_az","RFStacker","regression","Spearman",0.625,0.002),
    ("ld50_zhu","BaseBoosting","regression","MAE",0.552,0.009),
    ("ld50_zhu","ADMETrix","regression","MAE",0.573,0.010),
    ("ld50_zhu","MiniMol","regression","MAE",0.585,0.008),
    ("ld50_zhu","MACCS keys + autoML","regression","MAE",0.588,0.005),
    ("ld50_zhu","Chemprop","regression","MAE",0.606,0.024),
    ("herg","MapLight + GNN","binary","AUROC",0.880,0.002),
    ("herg","CFA","binary","AUROC",0.875,0.014),
    ("herg","SimGCN","binary","AUROC",0.874,0.014),
    ("herg","MapLight","binary","AUROC",0.871,0.004),
    ("herg","ZairaChem","binary","AUROC",0.856,0.009),
    ("ames","ZairaChem","binary","AUROC",0.871,0.002),
    ("ames","ADMETrix","binary","AUROC",0.870,0.006),
    ("ames","MapLight + GNN","binary","AUROC",0.869,0.002),
    ("ames","MapLight","binary","AUROC",0.868,0.002),
    ("ames","CFA","binary","AUROC",0.852,0.005),
    ("dili","MiniMol","binary","AUROC",0.956,0.006),
    ("dili","ZairaChem","binary","AUROC",0.925,0.005),
    ("dili","AttrMasking","binary","AUROC",0.919,0.008),
    ("dili","CFA","binary","AUROC",0.919,0.014),
    ("dili","MapLight + GNN","binary","AUROC",0.917,0.005),
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load_best_hyperparams(logs_dir, model_name, task):
    """
    Pull the best hyperparameter combo for a given model+task.
    Priority:
      1. <model>/logs/<task>_tuning.json  → "best_params" field
      2. <model>/artifacts/<task>/hyperparams.json → "params" field
    Returns a compact string, or "-" if nothing found.
    """
    tuning_path = os.path.join(logs_dir, model_name, "logs", f"{task}_tuning.json")
    if os.path.exists(tuning_path):
        try:
            with open(tuning_path) as f:
                tuning = json.load(f)
            best = tuning.get("best_params")
            if best:
                # BasicML/CaliciBoost-style: also carries a "best_model" / "best_combo" label
                label = tuning.get("best_model") or tuning.get("best_combo")
                parts = ", ".join(f"{k}={v}" for k, v in best.items())
                return f"{label}: {parts}" if label else parts
        except Exception:
            pass

    artifact_path = os.path.join(logs_dir, model_name, "artifacts", task, "hyperparams.json")
    if os.path.exists(artifact_path):
        try:
            with open(artifact_path) as f:
                hp = json.load(f)
            params = hp.get("params")
            if params:
                algo = hp.get("algorithm")
                parts = ", ".join(f"{k}={v}" for k, v in params.items())
                return f"{algo}: {parts}" if algo else parts
        except Exception:
            pass

    return "-"


def load_all_results(logs_dir):
    records = {}
    for model_name in MODELS:
        model_log_dir = os.path.join(logs_dir, model_name, "logs")
        if not os.path.exists(model_log_dir):
            continue
        log_files = glob(os.path.join(model_log_dir, "*.json"))
        log_files = [f for f in log_files
                     if not os.path.basename(f).startswith("_")
                     and "_tuning" not in os.path.basename(f)
                     and "_error"  not in os.path.basename(f)]
        for log_file in log_files:
            with open(log_file) as f:
                result = json.load(f)
            if "error" in result or result.get("status") == "skipped":
                continue
            task = result.get("task")
            if not task:
                continue
            if task not in records:
                records[task] = {}
            records[task][model_name] = {
                "score_mean":      result.get("score_mean"),
                "score_std":       result.get("score_std"),
                "metric":          result.get("metric", ""),
                "algorithm":       result.get("algorithm", MODEL_ALGORITHM.get(model_name, "-")),
                "best_hyperparams": _load_best_hyperparams(logs_dir, model_name, task),
            }
    return records


def style_header(ws, row, cols, bg_color="1F4E79", font_color="FFFFFF"):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(bold=True, color=font_color, name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=bg_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_row(ws, row, cols, bg_color=None):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if bg_color:
            cell.fill = PatternFill("solid", start_color=bg_color)


# ── Sheet builders ────────────────────────────────────────────────────────────

def add_leaderboard_sheet(wb):
    ws = wb.create_sheet("TDC_Leaderboard")
    headers = ["benchmark", "rank", "model", "task_type", "leaderboard_metric", "score", "std"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for i, (task, model, task_type, metric, score, std) in enumerate(TDC_LEADERBOARD):
        rank = (i % 5) + 1
        ws.append([task, rank, model, task_type, metric, score, std])
        bg = "D9E1F2" if rank % 2 == 0 else None
        style_row(ws, ws.max_row, len(headers), bg_color=bg)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 8


def add_blocked_sheet(wb, model_name):
    """Add a sheet for models that could not be run."""
    ws = wb.create_sheet(model_name)
    title, reasons = BLOCKED_MODELS[model_name]

    ws["A1"] = f"❌ {title}"
    ws["A1"].font = Font(bold=True, color="FF0000", name="Arial", size=12)

    for i, line in enumerate(reasons, start=2):
        ws.cell(row=i, column=1).value = line
        ws.cell(row=i, column=1).font  = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 90


def add_model_sheet(wb, model_name, records):
    """Add a results sheet for a model that ran successfully."""

    if model_name in BLOCKED_MODELS:
        add_blocked_sheet(wb, model_name)
        return

    ws = wb.create_sheet(model_name)

    # For CaliciBoost: note partial coverage at top
    if model_name in MODEL_TASK_COVERAGE:
        covered = MODEL_TASK_COVERAGE[model_name]
        ws["A1"] = (f"Note: {model_name} only covers {len(covered)} task(s): "
                    f"{', '.join(sorted(covered))}. "
                    f"All other tasks are marked N/A.")
        ws["A1"].font = Font(bold=True, color="7F6000", name="Arial", size=10)
        ws.append([])  # blank row
        start_row = 3
    else:
        start_row = 1

    headers = ["benchmark", "model", "algorithm", "task_type",
               "leaderboard_metric", "score", "std", "best_hyperparams"]
    ws.append(headers)
    style_header(ws, ws.max_row, len(headers))

    # Determine which tasks to show for this model
    task_coverage = MODEL_TASK_COVERAGE.get(model_name, None)

    for i, task in enumerate(TASKS):
        # Skip tasks outside this model's coverage
        if task_coverage is not None and task not in task_coverage:
            ws.append([task, model_name,
                       MODEL_ALGORITHM.get(model_name, "-"),
                       TASK_TYPE.get(task, "-"),
                       "N/A", "N/A", "N/A", "N/A"])
            style_row(ws, ws.max_row, len(headers),
                      bg_color="F2F2F2" if i % 2 == 0 else "E0E0E0")
            continue

        task_data = records.get(task, {}).get(model_name)
        if task_data and task_data.get("score_mean") is not None:
            score       = round(task_data["score_mean"], 4)
            std         = round(task_data["score_std"],  4)
            metric      = task_data["metric"]
            algorithm   = task_data.get("algorithm", MODEL_ALGORITHM.get(model_name, "-"))
            hyperparams = task_data.get("best_hyperparams", "-")
        else:
            score       = "pending" if model_name not in BLOCKED_MODELS else "-"
            std         = "-"
            metric      = "-"
            algorithm   = MODEL_ALGORITHM.get(model_name, "-")
            hyperparams = "-"

        task_type = TASK_TYPE.get(task, "-")
        ws.append([task, model_name, algorithm, task_type, metric, score, std, hyperparams])
        bg = "D9E1F2" if i % 2 == 0 else None
        style_row(ws, ws.max_row, len(headers), bg_color=bg)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 45


# ── CSV comparison ────────────────────────────────────────────────────────────

def generate_csv(records, output_dir):
    # Only include runnable models in the comparison
    runnable_models = [m for m in MODELS if m not in BLOCKED_MODELS]

    rows = []
    for task in TASKS:
        task_records     = records.get(task, {})
        lower_is_better  = task in LOWER_IS_BETTER
        metric           = "-"

        # Get metric name from any available result
        for model in runnable_models:
            if model in task_records and task_records[model].get("metric"):
                metric = task_records[model]["metric"]
                break

        best_score  = None
        best_model  = None
        best_std    = None
        best_algo   = None
        best_params = None

        for model in runnable_models:
            # Skip if model doesn't cover this task
            coverage = MODEL_TASK_COVERAGE.get(model)
            if coverage is not None and task not in coverage:
                continue
            if model not in task_records:
                continue
            mean = task_records[model].get("score_mean")
            std  = task_records[model].get("score_std")
            if mean is None:
                continue
            is_better = (
                best_score is None or
                (lower_is_better     and mean < best_score) or
                (not lower_is_better and mean > best_score)
            )
            if is_better:
                best_score  = mean
                best_model  = model
                best_std    = std
                best_algo   = task_records[model].get(
                    "algorithm", MODEL_ALGORITHM.get(model, "-")
                )
                best_params = task_records[model].get("best_hyperparams", "-")

        rows.append({
            "benchmark":          task,
            "best_model":         best_model if best_model else "-",
            "algorithm":          best_algo  if best_algo  else "-",
            "task_type":          TASK_TYPE.get(task, "-"),
            "leaderboard_metric": metric,
            "score":              round(best_score, 4) if best_score is not None else "-",
            "std":                round(best_std,   4) if best_std   is not None else "-",
            "best_hyperparams":   best_params if best_params else "-",
        })

    df = pd.DataFrame(rows)
    out_path = os.path.join(output_dir, "benchmark_comparison.csv")
    df.to_csv(out_path, index=False)
    print(f"✓ Saved CSV: {out_path}")

    wins = df[df["best_model"] != "-"]["best_model"].value_counts()
    print(f"\nBest model per task (among completed runs):")
    for model, count in wins.items():
        print(f"  {model}: {count} task(s)")
    return df


def generate_all_results_csv(records, output_dir):
    """
    benchmark_results.csv — all models × all tasks in one flat table.
    One row per (task, model) combination.
    """
    runnable_models = [m for m in MODELS if m not in BLOCKED_MODELS]
    rows = []

    for task in TASKS:
        task_records = records.get(task, {})
        for model in runnable_models:
            coverage = MODEL_TASK_COVERAGE.get(model)
            if coverage is not None and task not in coverage:
                continue

            task_data = task_records.get(model)
            if task_data and task_data.get("score_mean") is not None:
                score      = round(task_data["score_mean"], 4)
                std        = round(task_data["score_std"],  4)
                metric     = task_data["metric"]
                algorithm  = task_data.get("algorithm", MODEL_ALGORITHM.get(model, "-"))
                status     = "complete"
                hyperparams = task_data.get("best_hyperparams", "-")
            else:
                score      = "-"
                std        = "-"
                metric     = "-"
                algorithm  = MODEL_ALGORITHM.get(model, "-")
                status     = "pending"
                hyperparams = "-"

            rows.append({
                "benchmark":          task,
                "model":              model,
                "algorithm":          algorithm,
                "task_type":          TASK_TYPE.get(task, "-"),
                "leaderboard_metric": metric,
                "score":              score,
                "std":                std,
                "best_hyperparams":   hyperparams,
                "status":             status,
            })

    df = pd.DataFrame(rows)
    out_path = os.path.join(output_dir, "benchmark_results.csv")
    df.to_csv(out_path, index=False)
    print(f"✓ Saved CSV: {out_path}")
    return df


def generate_model_csvs(records, output_dir):
    """
    Per-model CSVs — one file per model, saved to output_dir/model_csvs/.
    Each file has one row per task the model covers.
    """
    csv_dir = os.path.join(output_dir, "model_csvs")
    os.makedirs(csv_dir, exist_ok=True)

    runnable_models = [m for m in MODELS if m not in BLOCKED_MODELS]

    for model in runnable_models:
        rows = []
        coverage = MODEL_TASK_COVERAGE.get(model)

        for task in TASKS:
            # Skip tasks outside model's coverage
            if coverage is not None and task not in coverage:
                continue

            task_data = records.get(task, {}).get(model)
            if task_data and task_data.get("score_mean") is not None:
                score      = round(task_data["score_mean"], 4)
                std        = round(task_data["score_std"],  4)
                metric     = task_data["metric"]
                algorithm  = task_data.get("algorithm", MODEL_ALGORITHM.get(model, "-"))
                status     = "complete"
                hyperparams = task_data.get("best_hyperparams", "-")
            else:
                score      = "-"
                std        = "-"
                metric     = "-"
                algorithm  = MODEL_ALGORITHM.get(model, "-")
                status     = "pending"
                hyperparams = "-"

            rows.append({
                "benchmark":          task,
                "model":              model,
                "algorithm":          algorithm,
                "task_type":          TASK_TYPE.get(task, "-"),
                "leaderboard_metric": metric,
                "score":              score,
                "std":                std,
                "best_hyperparams":   hyperparams,
                "status":             status,
            })

        df = pd.DataFrame(rows)
        out_path = os.path.join(csv_dir, f"{model}.csv")
        df.to_csv(out_path, index=False)
        print(f"✓ Saved CSV: {out_path}")

    # Also save blocked model stubs so every model has a file
    for model in BLOCKED_MODELS:
        df = pd.DataFrame([{
            "benchmark":          "-",
            "model":              model,
            "algorithm":          MODEL_ALGORITHM.get(model, "-"),
            "task_type":          "-",
            "leaderboard_metric": "-",
            "score":              "-",
            "std":                "-",
            "best_hyperparams":   "-",
            "status":             "could_not_run",
        }])
        out_path = os.path.join(csv_dir, f"{model}.csv")
        df.to_csv(out_path, index=False)
        print(f"✓ Saved CSV: {out_path} (blocked — could not run)")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--logs",   type=str, default="./model_assets")
    parser.add_argument("--output", type=str, default="./results")
    args = parser.parse_args()

    os.makedirs(args.output, exist_ok=True)

    print("Loading logs...")
    records = load_all_results(args.logs)
    print(f"Loaded results for {len(records)} tasks across "
          f"{sum(len(v) for v in records.values())} model-task pairs.")

    # ── CSVs ──────────────────────────────────────────────────────────────────
    print("\nGenerating benchmark_comparison.csv (best model per task)...")
    generate_csv(records, args.output)

    print("\nGenerating benchmark_results.csv (all models × all tasks)...")
    generate_all_results_csv(records, args.output)

    print("\nGenerating per-model CSVs (results/model_csvs/<model>.csv)...")
    generate_model_csvs(records, args.output)

    # ── Excel ─────────────────────────────────────────────────────────────────
    print("\nGenerating benchmark_results.xlsx...")
    wb = Workbook()
    wb.remove(wb.active)

    add_leaderboard_sheet(wb)
    for model in MODELS:
        add_model_sheet(wb, model, records)

    xlsx_path = os.path.join(args.output, "benchmark_results.xlsx")
    wb.save(xlsx_path)
    print(f"✓ Saved Excel: {xlsx_path}")
    print(f"\nSheets: TDC_Leaderboard, {', '.join(MODELS)}")

    print("\n✓ All outputs saved to:", args.output)
    print("  benchmark_comparison.csv  — best model per task")
    print("  benchmark_results.csv     — all models × all tasks")
    print("  model_csvs/<model>.csv    — one file per model")
    print("  benchmark_results.xlsx    — Excel workbook with all sheets")